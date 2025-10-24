# services/report_service.py
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from db import models
from typing import List
from datetime import datetime

# --- Constantes y _format_row (sin cambios) ---
HEADERS = [
    "OT", "Sitio", "Fecha y Hora", "Gls. Ex.",
    "Gls. Abast.", "Horometraje", "Estado", "Alarmas"
]

def _format_row(ab: models.Abastecimiento) -> list:
    fecha_str = ab.fecha.strftime('%d/%m/%y %H:%M') if isinstance(ab.fecha, datetime) else str(ab.fecha)
    sitio_nombre = ab.sitio.nombre if ab.sitio else "N/A"
    estado = str(ab.status.value) if hasattr(ab.status, 'value') else str(ab.status)
    alarmas_list = []
    if ab.alarma_transferencia: alarmas_list.append("Transferencia")
    if ab.alarma_falla_energia: alarmas_list.append("Falla Energía")
    alarmas_str = ", ".join(alarmas_list) if alarmas_list else "Sin Alarmas"
    return [
        ab.ot, sitio_nombre, fecha_str, f"{ab.gls_existentes:.2f}",
        f"{ab.gls_abastecidos:.2f}", f"{ab.horometraje:.2f}", estado, alarmas_str
    ]

# --- Generador de PDF (CORREGIDO - Anchos Variables + Centrado) ---
class PDFWithHeader(FPDF):
    def header(self):
        # Logo
        try:
             self.image("https://storage.googleapis.com/umg2025/logo.png", x=10, y=8, w=33)
        except Exception as e:
            print(f"Advertencia: No se pudo cargar el logo. {e}")
            self.set_xy(10, 8)
            self.set_font('Arial', 'B', 10)
            self.cell(33, 10, 'Tigo', 0, 0, 'L')
        # Título
        self.set_xy(0, 15)
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, "Historial de Abastecimientos", 0, 1, 'C')
        self.ln(10)

    def footer(self):
        # Pie de página
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, 0, 'C')
        self.set_x(-50)
        self.cell(0, 10, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 0, 'R')


def generate_historial_pdf(historial_data: List[models.Abastecimiento]) -> bytes:
    pdf = PDFWithHeader()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='L') # Paisaje
    pdf.set_font("Arial", size=9)

    page_width = pdf.w - 2 * pdf.l_margin
    margin = pdf.l_margin

    # --- CORRECCIÓN: Volver a anchos variables ---
    # Ajusta estos valores según veas necesario para tu contenido
    col_widths = [30, 75, 35, 20, 20, 25, 20, 45] # Dar más espacio a Sitio y Alarmas
    if len(col_widths) != len(HEADERS): # Fallback (poco probable ahora)
        width_per_col = page_width / len(HEADERS)
        col_widths = [width_per_col] * len(HEADERS)
    # --- FIN CORRECCIÓN ---

    # --- Calcular X inicial para centrar la tabla (con los anchos variables) ---
    table_total_width = sum(col_widths)
    # Asegurarse que la tabla no sea más ancha que la página
    if table_total_width > page_width:
        # Reducir proporcionalmente si excede (o ajustar manualmente los anchos)
        scale_factor = page_width / table_total_width
        col_widths = [w * scale_factor for w in col_widths]
        table_total_width = sum(col_widths) # Recalcular
        print("Advertencia: Anchos de columna reducidos para caber en la página.")

    start_x = (pdf.w - table_total_width) / 2 # Calcular X para centrar
    # --- FIN CÁLCULO ---


    # Encabezados de tabla (sin cambios en la lógica de dibujo)
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(220, 220, 220)
    line_height_header = pdf.font_size * 1.5
    current_y_header = pdf.get_y()
    for i, header in enumerate(HEADERS):
        current_x_header = start_x + sum(col_widths[:i])
        pdf.set_xy(current_x_header, current_y_header)
        pdf.multi_cell(col_widths[i], line_height_header, header, border=1, align='C', fill=True)
    pdf.set_y(current_y_header + line_height_header)

    # Filas de datos (sin cambios en la lógica de dibujo con MultiCell y altura dinámica)
    pdf.set_font("Arial", size=8)
    pdf.set_fill_color(245, 245, 245)
    fill = False
    line_height_data = pdf.font_size * 1.8 # Espaciado vertical

    for ab in historial_data:
        # Control de salto de página
        if pdf.get_y() + line_height_data > pdf.page_break_trigger - pdf.b_margin:
            pdf.add_page(orientation='L')
            # Redibujar encabezados
            pdf.set_font("Arial", 'B', 9)
            pdf.set_fill_color(220, 220, 220)
            current_y_header = pdf.get_y()
            for i, header in enumerate(HEADERS):
                current_x_header = start_x + sum(col_widths[:i])
                pdf.set_xy(current_x_header, current_y_header)
                pdf.multi_cell(col_widths[i], line_height_header, header, border=1, align='C', fill=True)
            pdf.set_y(current_y_header + line_height_header)
            pdf.set_font("Arial", size=8) # Volver a fuente de datos

        row_data = _format_row(ab)
        start_y_row = pdf.get_y()
        max_h = line_height_data

        # Calcular altura máxima
        temp_pdf = FPDF()
        temp_pdf.add_page(orientation='L')
        temp_pdf.set_font("Arial", size=8)
        calculated_heights = []
        for i, item in enumerate(row_data):
            # Usar un margen interno pequeño (c_margin) para el cálculo
            effective_width = col_widths[i] - (2 * pdf.c_margin) if col_widths[i] > (2 * pdf.c_margin) else col_widths[i]
            lines = temp_pdf.multi_cell(effective_width, line_height_data / 1.5, str(item), split_only=True)
            calculated_heights.append(len(lines) * line_height_data)
        row_height = max(max_h, max(calculated_heights) if calculated_heights else max_h)

        # Dibujar celdas
        current_x_row = start_x
        for i, item in enumerate(row_data):
            align = 'L' if i in [0, 1, 7] else 'C' # Alinear OT, Sitio y Alarmas a la izquierda
            pdf.set_xy(current_x_row, start_y_row)
            pdf.multi_cell(col_widths[i], row_height, str(item), border=1, ln=3,
                           align=align, fill=fill)
            current_x_row += col_widths[i]

        pdf.set_y(start_y_row + row_height)
        fill = not fill

    pdf_output: bytes = pdf.output(dest='S')
    return pdf_output


# --- Generador de Excel (sin cambios) ---
def generate_historial_excel(historial_data: List[models.Abastecimiento]) -> bytes:
    # ... (código sin cambios) ...
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historial Abastecimientos"
    sheet.append(HEADERS)
    for cell in sheet[1]: cell.font = Font(bold=True)
    for ab in historial_data:
        row_data = _format_row(ab)
        formatted_row = []
        for i, item in enumerate(row_data):
            try:
                if i in [3, 4, 5]: formatted_row.append(float(item))
                else: formatted_row.append(item)
            except ValueError: formatted_row.append(item)
        sheet.append(formatted_row)
    column_letters = [chr(ord('A') + i) for i in range(len(HEADERS))]
    for col_idx, column_letter in enumerate(column_letters, 1):
        max_length = 0
        try:
            column = sheet[column_letter]
            for cell in column:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length: max_length = cell_len
                except: pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        except KeyError: print(f"Advertencia: Columna {column_letter} no encontrada.")
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()